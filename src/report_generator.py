import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, IconSetRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os

def generate_report(input_file, output_file):
    print("Loading data...")
    try:
        df = pd.read_csv(input_file)
    except FileNotFoundError:
        print(f"Error: File {input_file} not found.")
        return

    # --- Data Enrichment (Calculated Columns) ---
    print("Enriching data...")
    
    # 1. Total Sales
    df['TotalSales'] = df['Quantity'] * df['UnitPrice']
    
    # 2. Commission (Simulated 5%)
    df['Commission'] = df['TotalSales'] * 0.05
    
    # 3. Performance Rating
    # High: > 10M, Medium: > 5M, Low: <= 5M
    def rate_performance(sales):
        if sales > 10000000: return "High"
        elif sales > 5000000: return "Medium"
        else: return "Low"
        
    df['Performance'] = df['TotalSales'].apply(rate_performance)

    # Reorder columns for better flow
    cols = ['Date', 'Region', 'Category', 'Product', 'Quantity', 'UnitPrice', 'TotalSales', 'Commission', 'Performance']
    df = df[cols]

    # --- Excel Export ---
    print("Exporting to Excel...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sales Data', index=False)

    # --- Advanced Formatting (Smart Table) ---
    print("Applying Smart Table formatting...")
    wb = load_workbook(output_file)
    ws = wb['Sales Data']
    
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)
    
    # 1. Convert to Official Excel Table (Olive Green)
    tab = Table(displayName="SalesTable", ref=f"A1:{last_col_letter}{max_row}")
    # TableStyleMedium3 is typically the Green/Olive style
    style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # 2. Column Widths & Alignment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 6) # Increased padding for formatted numbers
        ws.column_dimensions[column].width = adjusted_width
        
    # Center align specific columns
    for col_name in ['Region', 'Category', 'Performance', 'Quantity']:
        # Find column index
        for cell in ws[1]:
            if cell.value == col_name:
                col_idx = cell.column
                for row in range(2, max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')

    # 3. Currency Formatting
    currency_cols = ['UnitPrice', 'TotalSales', 'Commission']
    for col_name in currency_cols:
        for cell in ws[1]:
            if cell.value == col_name:
                col_idx = cell.column
                for row in range(2, max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = '#,##0 "IDR"'

    # 4. Freeze Top Row
    ws.freeze_panes = 'A2'

    wb.save(output_file)
    print(f"Smart Table generated successfully: {output_file}")
    
    # --- 5. Summary Sheet (Important Tables) ---
    print("Creating Summary Sheet...")
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Calculate Aggregations
    # 1. Sales by Region
    sales_by_region = df.groupby('Region')[['TotalSales']].sum().sort_values('TotalSales', ascending=False)
    sales_by_region.reset_index(inplace=True)
    
    # 2. Sales by Category
    sales_by_category = df.groupby('Category')[['TotalSales']].sum().sort_values('TotalSales', ascending=False)
    sales_by_category.reset_index(inplace=True)
    
    # 3. Top 5 Products
    top_products = df.groupby('Product')[['Quantity', 'TotalSales']].sum().sort_values('Quantity', ascending=False).head(5)
    top_products.reset_index(inplace=True)

    # Create Summary Sheet directly with openpyxl
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    ws_summary = wb.create_sheet("Summary")

    # Main Sheet Title
    ws_summary['B1'] = "EXECUTIVE SUMMARY"
    ws_summary['B1'].font = Font(bold=True, size=16, color="333333")

    # Titles for Tables
    ws_summary['B3'] = "Sales by Region"
    ws_summary['F3'] = "Sales by Category"
    
    title_font = Font(bold=True, size=12, color="004D40")
    for cell_ref in ['B3', 'F3']:
        ws_summary[cell_ref].font = title_font

    # Helper to write DF to sheet and format as table
    def write_and_format_table(ws, df, start_row, start_col, table_name, style="TableStyleMedium3", show_total=False):
        # Write Header and Data
        rows = dataframe_to_rows(df, index=False, header=True)
        
        row_count = 0
        for r_idx, row in enumerate(rows):
            row_count += 1
            for c_idx, value in enumerate(row):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
                
        num_cols = len(df.columns)
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(start_col + num_cols - 1)
        
        # Calculate final row for the table range
        # row_count includes Header + Data
        # If show_total is True, we add 1 more row
        total_row_offset = 1 if show_total else 0
        final_table_row = start_row + row_count - 1 + total_row_offset
        
        ref = f"{start_col_letter}{start_row}:{end_col_letter}{final_table_row}"
        
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False,
                                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        
        if show_total:
            tab.tableStyleInfo.showLastRow = True
            tab.totalsRowShown = True
            
            total_row_idx = start_row + row_count
            
            # Label "Total"
            ws.cell(row=total_row_idx, column=start_col).value = "Total"
            ws.cell(row=total_row_idx, column=start_col).font = Font(bold=True)
            
            # Sum Formula
            last_col_idx = start_col + num_cols - 1
            col_letter = get_column_letter(last_col_idx)
            data_start = start_row + 1
            data_end = start_row + row_count - 1
            
            ws.cell(row=total_row_idx, column=last_col_idx).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            ws.cell(row=total_row_idx, column=last_col_idx).number_format = '#,##0 "IDR"'
            ws.cell(row=total_row_idx, column=last_col_idx).font = Font(bold=True)

        ws.add_table(tab)
        
        # Formatting
        for r in range(start_row + 1, start_row + row_count): 
            ws.cell(row=r, column=start_col).alignment = Alignment(horizontal='center')
            
        last_col_idx = start_col + num_cols - 1
        for r in range(start_row + 1, start_row + row_count):
            ws.cell(row=r, column=last_col_idx).number_format = '#,##0 "IDR"'
            
        return row_count + total_row_offset

    # Create Tables
    # Table 1: Region (Header at Row 4, Col 2 -> B4)
    # Titles at Row 3.
    rows_region = write_and_format_table(ws_summary, sales_by_region, 4, 2, "RegionSummary", show_total=True)
    
    # Table 2: Category (Header at Row 4, Col 6 -> F4)
    rows_category = write_and_format_table(ws_summary, sales_by_category, 4, 6, "CategorySummary", show_total=True)
    
    # Calculate start row for next table (Dynamic Spacing)
    # Max rows used by top tables + spacing
    max_rows_used = max(rows_region, rows_category)
    next_start_row = 4 + max_rows_used + 3 # 3 rows gap
    
    # Table 3: Top Products
    ws_summary[f'B{next_start_row - 1}'] = "Top 5 Best Selling Products"
    ws_summary[f'B{next_start_row - 1}'].font = title_font
    
    write_and_format_table(ws_summary, top_products, next_start_row, 2, "TopProductsSummary", show_total=False)

    # Auto-width
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 6) # Increased padding
        ws_summary.column_dimensions[column].width = adjusted_width
        
    wb.save(output_file)
    print("Summary Sheet added successfully.")

    # --- 6. Generate Email Summary (Distribution) ---
    print("\nGenerating Email Draft...")
    
    total_revenue = df['TotalSales'].sum()
    
    # Top Region
    top_region = df.groupby('Region')['TotalSales'].sum().idxmax()
    top_region_val = df.groupby('Region')['TotalSales'].sum().max()
    
    # Best Product (by Qty)
    best_product = df.groupby('Product')['Quantity'].sum().idxmax()
    best_product_qty = df.groupby('Product')['Quantity'].sum().max()
    
    current_month = datetime.now().strftime("%B %Y")
    
    email_body = f"""
=============================================================
[DRAFT EMAIL - READY TO SEND]
Subject: Laporan Penjualan Bulanan - {current_month}

Halo Tim,

Berikut adalah ringkasan penjualan untuk bulan ini:

1. Total Omzet    : Rp {total_revenue:,.0f}
2. Wilayah Terbaik: {top_region} (Rp {top_region_val:,.0f})
3. Produk Terlaris: {best_product} ({best_product_qty} unit)

Laporan detail (Excel) sudah terlampir. 
Mohon diperiksa.

Terima kasih.
=============================================================
"""
    print(email_body)
